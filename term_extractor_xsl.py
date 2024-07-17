import os
import pickle
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from docx import Document
from natasha import Segmenter, NewsEmbedding, NewsMorphTagger, MorphVocab, Doc



# Load term dictionaries
term_dict_files = ['arts_terms',
                    'bio_terms',
                    'cs_terms',
                    'math_terms',
                    'mus_terms',
                    'phys_terms',
                    'rus_terms']
term_dicts = {}

for term_file in term_dict_files:
    with open(os.path.join(os.getcwd(), 'termdicts', term_file), 'rb') as f:
        term_dicts[term_file.split('_')[0]] = pickle.load(f)
from natasha import Segmenter, NewsEmbedding, NewsMorphTagger, MorphVocab, Doc
# Natasha NLP setup
segmenter = Segmenter()
emb = NewsEmbedding()
morph_tagger = NewsMorphTagger(emb)
morph_vocab = MorphVocab()

class NLP_analyzer:
    def __init__(self, text):
        self.text = text

    def lemmatize(self):
        doc = Doc(self.text)
        doc.segment(segmenter)
        doc.tag_morph(morph_tagger)
        for token in doc.tokens:
            token.lemmatize(morph_vocab)
        return [token.lemma for token in doc.tokens]

def TERM_EXTRACTOR(termdict, lemmas):
    termdict_lemmas = termdict.keys()
    terms = []  # идентифицированные термины
    ignoring_for_text = set()  # для хранения информации о токенах, которые определены как термины

    max_len = 6
    for length in range(max_len, 0, -1):
        for i in range(len(lemmas) - length + 1):
            if any(j in ignoring_for_text for j in range(i, i + length)):
                continue
            collocation_candidate = " ".join(lemmas[i:i + length])
            if collocation_candidate in termdict_lemmas:
                terms.append((collocation_candidate, termdict[collocation_candidate]))
                ignoring_for_text.update(range(i, i + length))

    return terms

def read_text_from_file(file_path):
    if file_path.endswith('.txt'):
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    elif file_path.endswith('.docx') or file_path.endswith('.DOCX') : #компьютер в руках филолога -- кошмар программиста
        doc = Document(file_path)
        return '\n'.join([para.text for para in doc.paragraphs])
    else:
        raise ValueError("Unsupported file type")

def process_textbook(textbook_path):
    text = read_text_from_file(textbook_path)
    lemmas = NLP_analyzer(text).lemmatize()
    
    term_occurrences = {}

    for term_key, termdict in term_dicts.items():
        extracted_terms = TERM_EXTRACTOR(termdict, lemmas)
        for term, term_form in extracted_terms:
            if term_form not in term_occurrences:
                term_occurrences[term_form] = set()
            term_occurrences[term_form].add(term_key)
    
    return term_occurrences


def save_to_excel(textbook_term_data, output_filename='term_results.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = "Term Results"

    # Write headers
    headers = ["Термин", "Учебник", "В скольких списках встречался"] + term_dict_files
    ws.append(headers)

    # Define fills
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Write data
    for textbook, terms in textbook_term_data.items():
        for term, dicts in terms.items():
            row = [term, textbook, len(dicts)]
            for term_file in term_dict_files:
                term_key = term_file.split('_')[0]
                value = 1 if term_key in dicts else 0
                cell = value
                row.append(cell)
            ws.append(row)
            for col_num in range(4, len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=col_num)
                if cell.value == 1:
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill

    # Save the workbook
    wb.save(output_filename)

def main():
    textbooks_dir = 'textbooks'
    textbooks_list = [f for f in os.listdir(textbooks_dir) if os.path.isfile(os.path.join(textbooks_dir, f))]
    textbook_term_data = {}

    for textbook in textbooks_list:
        print ('Processing now: '+ textbook)
        textbook_path = os.path.join(textbooks_dir, textbook)
        term_occurrences = process_textbook(textbook_path)
        filtered_term_occurrences = {term: dicts for term, dicts in term_occurrences.items() if len(dicts) >= 2}
        textbook_term_data[textbook] = filtered_term_occurrences

    save_to_excel(textbook_term_data)

    print("Term extraction completed.")

if __name__ == "__main__":
    main()
